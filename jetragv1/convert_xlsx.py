"""
convert_xlsx.py  (v2 – narrative format for better RAG retrieval)
Converts the Mo Ambulance MIS Excel report into human-readable narrative
text files optimised for semantic vector search.

Run on your Mac:
  python3 convert_xlsx.py
"""

import openpyxl
import os

EXCEL_PATH = "Mo Ambulance_MIS_21_26.xlsx"
OUTPUT_DIR = "converted_docs"

os.makedirs(OUTPUT_DIR, exist_ok=True)

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

def fmt(v):
    """Format a number as Indian Rupee with commas, or return string as-is."""
    if v is None:
        return "N/A"
    if isinstance(v, (int, float)):
        return f"Rs {v:,.0f}"
    return str(v).strip()

def pct(v):
    if v is None:
        return "N/A"
    try:
        return f"{float(v)*100:.1f}%"
    except:
        return str(v)

# ── 1. Assumptions & Definitions ───────────────────────────────────────────
ws = wb["Assumption & Definition"]
lines = [
    "Mo Ambulance MIS Report – Glossary and Definitions\n",
    "The following terms are used throughout the Mo Ambulance MIS report:\n"
]
for row in ws.iter_rows(values_only=True):
    vals = [v for v in row if v is not None and str(v).strip()]
    if len(vals) >= 2:
        term = str(vals[0]).strip()
        desc = str(vals[1]).strip()
        extra = str(vals[2]).strip() if len(vals) > 2 else ""
        if term not in ("Mo Ambulance", "A") and "DEFINITION" not in term:
            line = f"{term}: {desc}"
            if extra:
                line += f" — {extra}"
            lines.append(line)

with open(f"{OUTPUT_DIR}/01_glossary.txt", "w") as f:
    f.write("\n".join(lines))
print("✓ Created 01_glossary.txt")


# ── 2. Revenue Summary (narrative) ─────────────────────────────────────────
ws = wb["Revenue"]
all_rows = [row for row in ws.iter_rows(values_only=True)]

# Row layout from inspection:
# Row 4 = headers, Rows 5+ = data years
# Cols: B=year, C=AggRide, D=HAMS, E=Industry, F=MHU, G=OxygenSensor, H=IoT, I=Kiosk, J=WhiteLabel, K=Total, L=Expenditure, M=EBITDA, N=EBITDA%

year_data = []
for row in all_rows:
    if row[1] and str(row[1]).strip() in ["2021-22","2022-23","2023-24","2024-25","2025-26"]:
        year_data.append(row)

lines = [
    "Mo Ambulance – Five-Year Revenue and EBITDA Summary\n",
    "This section summarises Mo Ambulance's financial performance from FY 2021-22 to FY 2025-26.\n"
]

for row in year_data:
    year     = str(row[1]).strip() if row[1] else "?"
    rides    = row[2]
    hams     = row[3]
    industry = row[4]
    mhu      = row[5]
    oxygen   = row[6]
    iot      = row[7]
    kiosk    = row[8]
    white    = row[9]
    total    = row[10]
    expense  = row[11]
    ebitda   = row[12]
    ebitda_p = row[13]

    para = (
        f"In FY {year}, Mo Ambulance's total revenue was {fmt(total)} "
        f"against total expenditure of {fmt(expense)}, resulting in an EBITDA of {fmt(ebitda)} ({pct(ebitda_p)}). "
        f"Revenue breakdown: Aggregator Rides (B2C) contributed {fmt(rides)}, "
        f"HAMS (Hospital Ambulance Management System, B2B) contributed {fmt(hams)}, "
        f"Industry (B2B) contributed {fmt(industry)}, "
        f"MHU (Mobile Health Unit) contributed {fmt(mhu)}, "
        f"IoT/ATOMS contributed {fmt(iot)}, "
        f"Pre-paid Kiosks contributed {fmt(kiosk)}, "
        f"White Labelling contributed {fmt(white)}."
    )
    lines.append(para)
    lines.append("")

with open(f"{OUTPUT_DIR}/02_revenue_summary.txt", "w") as f:
    f.write("\n".join(lines))
print("✓ Created 02_revenue_summary.txt")


# ── 3. Expenses (narrative) ─────────────────────────────────────────────────
ws = wb["Expenses"]
all_rows = list(ws.iter_rows(values_only=True))

lines = [
    "Mo Ambulance – Detailed Expense Projections (FY 2021-22 to FY 2025-26)\n",
    "This section covers year-wise income and expenditure by expense head.\n"
]

header_found = False
headers = []
for row in all_rows:
    vals = [str(v).strip() if v is not None else "" for v in row]
    non_empty = [v for v in vals if v]
    if not non_empty:
        continue
    if "Expences Head" in non_empty or "Expenses Head" in non_empty:
        headers = vals
        header_found = True
        continue
    if header_found:
        expense_head = vals[1] if len(vals) > 1 and vals[1] else None
        if expense_head:
            data_parts = [f"{headers[i]}: {vals[i]}" for i in range(2, len(vals)) if i < len(headers) and vals[i] and headers[i]]
            if data_parts:
                lines.append(f"Expense Head '{expense_head}': " + " | ".join(data_parts))

with open(f"{OUTPUT_DIR}/03_expenses.txt", "w") as f:
    f.write("\n".join(lines))
print("✓ Created 03_expenses.txt")


# ── 4. Year-wise sheets ─────────────────────────────────────────────────────
YEAR_SHEETS = {
    "2021-22": ("04_detail_2021_22.txt", "FY 2021-22"),
    "2022-23": ("05_detail_2022_23.txt", "FY 2022-23"),
    "2023-24": ("06_detail_2023_24.txt", "FY 2023-24"),
    "2024-25": ("07_detail_2024_25.txt", "FY 2024-25"),
    "2025-26": ("08_detail_2025_26.txt", "FY 2025-26"),
}

for sheet_name, (filename, label) in YEAR_SHEETS.items():
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    
    lines = [
        f"Mo Ambulance – {label} Detailed Revenue by Business Line and State\n",
        f"This section provides a state-wise, business-line breakdown of all revenue for {label}.\n"
    ]
    
    current_source = None
    subtotal_label = None
    
    for row in all_rows:
        vals = [v for v in row]
        str_vals = [str(v).strip() if v is not None else "" for v in vals]
        non_empty = [v for v in str_vals if v]
        
        if not non_empty:
            continue

        # Detect revenue source (col index 1)
        if len(vals) > 1 and vals[1] and str(vals[1]).strip() and str(vals[1]).strip() not in ["Revenue Sources", "Revenue Sources "]:
            candidate = str(vals[1]).strip()
            if not candidate[0].isdigit():
                current_source = candidate

        # Detect total rows
        if "Total" in str_vals or "total" in str_vals or "otal" in " ".join(non_empty):
            # Find the revenue figure (last non-empty numeric value)
            total_val = None
            for v in reversed(vals):
                if isinstance(v, (int, float)):
                    total_val = v
                    break
            if total_val and current_source:
                lines.append(f"Total revenue from {current_source} in {label}: {fmt(total_val)}.")
            elif total_val:
                lines.append(f"Grand total revenue for {label}: {fmt(total_val)}.")
            continue

        # Detect data rows with state info (col index 2)
        if len(vals) > 2 and vals[2] and str(vals[2]).strip() and current_source:
            state = str(vals[2]).strip()
            # Skip header-like states
            if state in ["States / Cities", "Revenue Sources", "States"]:
                continue
            
            # Find the revenue figure (last column with a numeric value)
            revenue = None
            for v in reversed(vals):
                if isinstance(v, (int, float)) and v > 100:  # filter out counts/ratios
                    revenue = v
                    break
            
            # Get relevant numeric data for context
            numbers = [(i, v) for i, v in enumerate(vals) if isinstance(v, (int, float)) and v > 0]
            
            if revenue and revenue > 1000:
                detail = f"In {label}, {current_source} in {state} generated revenue of {fmt(revenue)}."
                # Add unit counts if present
                for idx, num in numbers:
                    if num != revenue and num > 0 and num < 10000:
                        detail += f" (units/count: {num:.0f})"
                        break
                lines.append(detail)
        
    with open(f"{OUTPUT_DIR}/{filename}", "w") as f:
        f.write("\n".join(lines))
    print(f"✓ Created {filename}")


print(f"\n✅ All narrative files written to ./{OUTPUT_DIR}/")
